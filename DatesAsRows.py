import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import textwrap
from openpyxl.utils import get_column_letter



class TimesheetContext:
    """Data class to hold shared context information."""
    def __init__(self, df):
        self.df = df
        self.df_name = df.iloc[1, 1]     # B2
        self.df_actual_date = df.iloc[2, 35]  # AJ3
        self.df_client_signature = df.iloc[34, 0]  # A35
        self.df_consultant_signature = df.iloc[34, 34]  # AI35
        self.df_client_signature_date = df.iloc[34, 36]  # AK35
        self.df_consultant_signature_date = df.iloc[34, 5]  # E35
        self.df_days_of_month = df.iloc[6, 1:32]  # Row 7 (B7:AF7)
        self.df_days_of_week = df.iloc[7, 1:32]  # Row 8 (B8:AF8)


class TimesheetProcessor:
    """Class to process timesheets and generate reports."""

    def __init__(self):
        self.thin_side = Side(style='thin')
        self.bold_font = Font(bold=True)

    @staticmethod
    def extract_category_components(category_str):
        pattern = r"^\s*(Billable|Unbillable)\s*-\s*(.+?)\s*-\s*(.+?)\s*$"
        match = re.match(pattern, str(category_str).strip(), re.IGNORECASE)
        if match:
            category_type = match.group(1).strip()
            client = match.group(2).strip()
            project = match.group(3).strip()
            return category_type, client, project
        else:
            return None, None, None
        



    def process_category(self, df, category_type, client, project):
        # Filter rows for this client-project pair and category type
        df_filtered = df[
            (df['CategoryType'].str.casefold() == category_type.casefold()) &
            (df['Client'].str.casefold() == client.casefold()) &
            (df['Project'].str.casefold() == project.casefold())
        ]

        if df_filtered.empty:
            return pd.DataFrame({'Hours': [0]*31, 'Days': [0]*31}), 0, 0

        # Select columns 1 to 31 (dates)
        data = df_filtered.iloc[:, 1:32].apply(pd.to_numeric, errors='coerce')

        # Sum the data per date (column-wise)
        summed_data = data.sum(axis=0).reset_index(drop=True)

        # Create DataFrame
        df_summed = pd.DataFrame(summed_data, columns=['Hours'])

        # Calculate Days
        df_summed['Days'] = (df_summed['Hours'] / 7.5).round(2)

        # Calculate totals
        total_hours = df_summed['Hours'].sum()
        total_days = df_summed['Days'].sum()

        return df_summed, total_hours, total_days

    def collect_notes(self, df, client, project):
        # Filter rows for this client-project pair
        df_filtered = df[
            (df['Client'].str.casefold() == client.casefold()) &
            (df['Project'].str.casefold() == project.casefold())
        ]

        # Initialize list of sets to store unique notes per date
        accumulated_notes = [set() for _ in range(31)]  # Assuming 31 days max

        # For each row in df_filtered
        for _, row in df_filtered.iterrows():
            note = str(row.iloc[34]).strip()
            if note == '' or pd.isna(note):
                continue
            for col_idx in range(1, 32):  # Columns B to AF (1 to 31)
                value = row.iloc[col_idx]
                if pd.notna(value):
                    date_index = col_idx - 1  # Adjust index to 0-based
                    accumulated_notes[date_index].add(note)

        # Convert sets to concatenated strings
        notes_list = ['; '.join(sorted(notes)) if notes else '' for notes in accumulated_notes]

        # Return DataFrame of notes
        notes_df = pd.DataFrame(notes_list, columns=['Notes'])
        return notes_df

    def setup_worksheet(self, ws):

        # Insert a row before row 4
        ws.insert_rows(4)

        # Insert static data into the sheet
        ws["A1"] = "Employee and Consultant Timesheet"
        ws["A3"] = "Client:"
        ws["A4"] = "DATE"
        ws["H3"] = "Period:"
        ws["H2"] = "Project ID:"
        ws["A40"] = "Consultant Signature"
        ws["C4"] = "Billable"
        ws["E4"] = "Unbillable"
        ws["C5"] = "Hours"
        ws["D5"] = "Days"
        ws["E5"] = "Hours"
        ws["F5"] = "Days"
        ws["A2"] = "Name:"
        ws["A42"] = (
            "E-Mail signed timesheets and invoices to ap@closereach.ca by "
            "1st business day of following billing period."
        )
        ws["G42"] = (
            "Signature by Client indicates acceptance of billable time and "
            "satisfaction with work performed."
        )
        ws["E40"] = "Date"
        ws["I40"] = "Date"
        ws["G40"] = "Client Signature"
        ws["G4"] = "Notes/Status Information/Project Codes"

        # Set bold fonts for headers
        ws["A2"].font = self.bold_font
        ws["A1"].font = self.bold_font
        ws["A1"].alignment = Alignment(horizontal='center')
        ws["A3"].font = self.bold_font
        ws["D3"].font = self.bold_font
        ws["D2"].font = self.bold_font
        ws["A40"].font = self.bold_font
        ws["G4"].font = self.bold_font
        ws["H2"].font = self.bold_font
        ws["H3"].font = self.bold_font
        ws["I40"].font = self.bold_font
        ws["G40"].font = self.bold_font
        ws["E40"].font = self.bold_font
        ws["H3"].font = self.bold_font

        # Apply underline to cell B2 (Excel indexing starts at 1, so B2 is row 2, column 2)
        ws['B2'].font = Font(underline='single')
        ws['B3'].font = Font(underline='single')
        ws['I2'].font = Font(underline='single')
        ws['I3'].font = Font(underline='single')


        # Center align the cells from A2 to I3
        for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=9):  # A=1, I=9
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        # List of cells that should be center-aligned
        center_cells = ['A4', 'C4', 'C5', 'D5', 'E4', 'E5', 'F5']

        # Loop through each cell and apply center alignment
        for cell_ref in center_cells:
            ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')


        # Merge header cells G4:I4
        ws.merge_cells('G4:I4')
        ws["G4"].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Merge columns G, H, and I from row 6 to row 37
        for row in range(6, 38):  # Rows 6 to 37 inclusive
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
            # Set alignment and wrap_text for the merged cell
            cell = ws.cell(row=row, column=7)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


        # Merge cells A42 to D46
        ws.merge_cells('A42:D46')
        # Apply center alignment and wrap text
        ws["A42"].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Merge cells A42 to D46
        ws.merge_cells('G42:H46')
        # Apply center alignment and wrap text
        ws["G42"].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Merge cells A42 to D46
        ws.merge_cells('A39:D39')


        # Merge cells A42 to D46
        ws.merge_cells('G39:H39')


        from openpyxl.styles import Border, Side

        # Define a thin border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply the thin border to each cell in the range A6:I37
        for row in ws.iter_rows(min_row=6, max_row=37, min_col=1, max_col=9):  # A=1, I=9
            for cell in row:
                cell.border = thin_border

        # Define a bold font
        bold_font = Font(bold=True)

        # Apply the bold font to each cell in the range A6:B36
        for row in ws.iter_rows(min_row=6, max_row=36, min_col=1, max_col=2):  # A=1, B=2
            for cell in row:
                cell.font = bold_font

    def insert_data_into_worksheet(self, ws, context, billable_data, unbillable_data, notes_df, client, project):
        num_days = len(context.df_days_of_month)

        # Calculate the total width of merged columns G, H, and I
        merged_columns = ['G', 'H', 'I']
        total_width = sum([ws.column_dimensions[col].width if ws.column_dimensions[col].width else 10 for col in merged_columns])


        # List of scattered cells that should be center-aligned
        scattered_cells = ['E40', 'I40']

        # Loop through each cell and apply center alignment
        for cell_ref in scattered_cells:
            ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')
        
        # Estimate the number of characters that fit in the merged cell based on total width
        # The multiplier 1.2 adjusts for average character width; you might need to tweak this
        average_char_width = 1.2  
        max_chars_per_line = int(total_width / average_char_width)

        for r_idx in range(num_days):
            row_num = r_idx + 6  # Start from row 5

            date = context.df_days_of_month.iloc[r_idx]
            day_of_week = context.df_days_of_week.iloc[r_idx]

            ws.cell(row=row_num, column=1, value=date)  # DATE in column A
            ws.cell(row=row_num, column=2, value=day_of_week)  # Day of Week in column B

            # Billable data
            billable_hours = billable_data['Hours'].iloc[r_idx] if r_idx < len(billable_data) else 0
            billable_days = billable_data['Days'].iloc[r_idx] if r_idx < len(billable_data) else 0
            ws.cell(row=row_num, column=3, value=billable_hours)  # Billable Hours in column C
            ws.cell(row=row_num, column=4, value=billable_days)  # Billable Days in column D

            # Unbillable data
            unbillable_hours = unbillable_data['Hours'].iloc[r_idx] if r_idx < len(unbillable_data) else 0
            unbillable_days = unbillable_data['Days'].iloc[r_idx] if r_idx < len(unbillable_data) else 0
            ws.cell(row=row_num, column=5, value=unbillable_hours)  # Unbillable Hours in column E
            ws.cell(row=row_num, column=6, value=unbillable_days)  # Unbillable Days in column F

            # Notes
            note = notes_df['Notes'].iloc[r_idx] if r_idx < len(notes_df) else ''
            cell = ws.cell(row=row_num, column=7, value=note)  # Notes in column G (merged with H and I)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # Calculate the number of lines required for the note
            if note:
                # Split the note into existing lines
                existing_lines = note.split('\n')
                total_wrapped_lines = 0
                for line in existing_lines:
                    # Use textwrap to wrap each line based on max_chars_per_line
                    wrapped = textwrap.wrap(line, width=max_chars_per_line)
                    if not wrapped:
                        # If the line is empty, count it as one line
                        total_wrapped_lines += 1
                    else:
                        total_wrapped_lines += len(wrapped)
                num_lines = total_wrapped_lines
                # Set the row height based on the number of lines
                # Adjust the multiplier (15) as needed for your font size
                ws.row_dimensions[row_num].height = num_lines * 15
            else:
                # Set to default row height
                ws.row_dimensions[row_num].height = 15

        # Insert totals
        total_row = num_days + 6  # Adjust total row based on number of days
        ws.cell(row=total_row, column=1, value="Total").font = self.bold_font
        ws.cell(row=total_row, column=3, value=billable_data['Hours'].sum()).font = self.bold_font
        ws.cell(row=total_row, column=4, value=billable_data['Days'].sum()).font = self.bold_font
        ws.cell(row=total_row, column=5, value=unbillable_data['Hours'].sum()).font = self.bold_font
        ws.cell(row=total_row, column=6, value=unbillable_data['Days'].sum()).font = self.bold_font

        # Insert additional data
        ws["B2"] = context.df_name
        ws["B3"] = client
        ws["I2"] = project
        ws["I3"] = context.df_actual_date  # Assuming df_actual_date represents the month
        ws["A39"] = context.df_consultant_signature
        ws["E39"] = context.df_consultant_signature_date
        ws["G39"] = context.df_client_signature
        ws["I39"] = context.df_client_signature_date




        # Merge cells from A2 to I3
        ws.merge_cells('B2:G2')
        ws.merge_cells('B3:G3')
        ws.merge_cells('A4:B5')
        ws.merge_cells('C4:D4')
        ws.merge_cells('E4:F4')
        ws.merge_cells('A1:I1')
        ws.merge_cells('A37:B37')
        ws.merge_cells('E39:F39')

        ws['I39'].font = Font(size=7)
        ws['E39'].font = Font(size=7)
        
        # List of cells that should be set to bold
        bold_cells = ['A4', 'C4', 'C5', 'D5', 'E4', 'E5', 'F5']

        # Loop through each cell and set it to bold
        for cell_ref in bold_cells:
            ws[cell_ref].font = Font(bold=True)

        # Loop through each cell in the range A6:F37 and apply center alignment
        for row in ws.iter_rows(min_row=6, max_row=37, min_col=1, max_col=6):  # A=1, F=6
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def process_client_project(self, wb, context, df, client, project):
        # Process billable category
        billable_data, _, _ = self.process_category(df, 'Billable', client, project)

        # Process unbillable category
        unbillable_data, _, _ = self.process_category(df, 'Unbillable', client, project)

        # Collect notes (from both billable and unbillable entries)
        notes_df = self.collect_notes(df, client, project)

        # Create sheet
        sheet_name = f"{client} - {project}"[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        # Set up the worksheet structure
        self.setup_worksheet(ws)

        # Insert data
        self.insert_data_into_worksheet(
            ws, context, billable_data, unbillable_data, notes_df, client, project
        )
    
        def set_column_widths(ws, column_widths):
            
            for col, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = width

        # Example usage
        column_widths = {1:7, 2:5, 3:6, 4:6,5:6,6:6}  # Columns 1 (A), 2 (B), 7 (G) with widths
        set_column_widths(ws, column_widths)


        # Adjust the widths of columns G, H, and I
        ws.column_dimensions['G'].width = 15  # Set the width of column G
        ws.column_dimensions['H'].width = 15  # Set the width of column H
        ws.column_dimensions['I'].width = 20  # Set the width of column I


        ## Define a thin border for the outer edges
        thin_side = Side(style='thin')

        # Apply border to the outer edges of the range A1:I3

        # Apply top border to all cells in the first row (A1:I1)
        for cell in ws["A1:I1"][0]:  # ws["A1:I1"] gives the rows, we select the first row (top)
            cell.border = Border(top=thin_side)

        # Apply bottom border to all cells in the last row (A3:I3)
        for cell in ws["A3:I3"][0]:  # ws["A3:I3"] gives the rows, we select the last row (bottom)
            cell.border = Border(bottom=thin_side)

        # Apply left border to all cells in the first column (A1:A3)
        for cell in ws["A1:A3"]:  # ws["A1:A3"] gives the columns, we select the first column (left)
            cell[0].border = Border(left=thin_side)

        # Apply right border to all cells in the last column (I1:I3)
        for cell in ws["I1:I3"]:  # ws["I1:I3"] gives the columns, we select the last column (right)
            cell[0].border = Border(right=thin_side)

        # Now handle the corner cells to add missing borders

        # Top-left corner (A1) - needs both top and left borders
        ws['A1'].border = Border(left=thin_side, top=thin_side)

        # Bottom-left corner (A3) - needs both bottom and left borders
        ws['A3'].border = Border(left=thin_side, bottom=thin_side)

        # Top-right corner (I1) - needs both top and right borders
        ws['I1'].border = Border(right=thin_side, top=thin_side)

        # Bottom-right corner (I3) - needs both bottom and right borders
        ws['I3'].border = Border(right=thin_side, bottom=thin_side)


        # Define a thin border for the outer edges
        thin_side = Side(style='thin')


        # Define a thin border for all sides
        thin_side = Side(style='thin')

        # Apply top border to all cells in the first row (A39:I39)
        for cell in ws["A39:I39"][0]:  # Select the first row of the range
            cell.border = Border(top=thin_side)

        # Apply bottom border to all cells in the last row (A46:I46)
        for cell in ws["A46:I46"][0]:  # Select the last row of the range
            cell.border = Border(bottom=thin_side)

        # Apply left border to all cells in the first column (A39:A46)
        for cell in ws["A39:A46"]:  # Select the first column of the range
            cell[0].border = Border(left=thin_side)

        # Apply right border to all cells in the last column (I39:I46)
        for cell in ws["I39:I46"]:  # Select the last column of the range
            cell[0].border = Border(right=thin_side)

        # Handle corner cells to ensure both top/bottom and left/right borders are applied
        ws['A39'].border = Border(left=thin_side, top=thin_side)      # Top-left corner (A39)
        ws['A46'].border = Border(left=thin_side, bottom=thin_side)   # Bottom-left corner (A46)
        ws['I39'].border = Border(right=thin_side, top=thin_side)     # Top-right corner (I39)
        ws['I46'].border = Border(right=thin_side, bottom=thin_side)  # Bottom-right corner (I46)



        # Define a thin border for the outer edges
        thin_side = Side(style='thin')

        # Apply border to the outer edges of the merged range A4:B5

        # Apply top border to cells in the first row (A4:B4)
        for cell in ws["A4:B4"][0]:  # First row of the merged range
            cell.border = Border(top=thin_side)

        # Apply bottom border to cells in the last row (A5:B5)
        for cell in ws["A5:B5"][0]:  # Last row of the merged range
            cell.border = Border(bottom=thin_side)

        # Apply left border to the first column in the range (A4:A5)
        for cell in ws["A4:A5"]:  # Left column of the merged range
            cell[0].border = Border(left=thin_side)

        # Apply right border to the last column in the range (B4:B5)
        for cell in ws["B4:B5"]:  # Right column of the merged range
            cell[0].border = Border(right=thin_side)

        # Ensure the corner cells have the necessary borders

        # Top-left corner (A4) - needs both top and left borders
        ws['A4'].border = Border(left=thin_side, top=thin_side)

        # Bottom-left corner (A5) - needs both bottom and left borders
        ws['A5'].border = Border(left=thin_side, bottom=thin_side)

        # Top-right corner (B4) - needs both top and right borders
        ws['B4'].border = Border(right=thin_side, top=thin_side)

        # Bottom-right corner (B5) - needs both bottom and right borders
        ws['B5'].border = Border(right=thin_side, bottom=thin_side)

        # Apply border to the outer edges of the merged range C4:D4

        # Apply top border to cells in the first row (C4:D4)
        for cell in ws["C4:D4"][0]:  # First row of the merged range
            cell.border = Border(top=thin_side)

        # Apply bottom border to cells in the same row (C4:D4) since it's only one row
        for cell in ws["C4:D4"][0]:  # First and last row of the merged range
            cell.border = Border(bottom=thin_side)

        # Apply left border to the first column in the range (C4:C4)
        ws["C4"].border = Border(left=thin_side)

        # Apply right border to the last column in the range (D4:D4)
        ws["D4"].border = Border(right=thin_side)

        # Ensure the corner cells have the necessary borders
        # Top-left corner (C4) - needs both top and left borders
        ws['C4'].border = Border(left=thin_side, top=thin_side)

        # Top-right corner (D4) - needs both top and right borders
        ws['D4'].border = Border(right=thin_side, top=thin_side)

        # Since this is a single row merge, no bottom-left and bottom-right corners needed separately

        # Apply border to the outer edges of the merged range E4:F4

        # Apply top border to cells in the first row (E4:F4)
        for cell in ws["E4:F4"][0]:  # First row of the merged range
            cell.border = Border(top=thin_side)

        # Apply bottom border to cells in the same row (E4:F4) since it's only one row
        for cell in ws["E4:F4"][0]:  # First and last row of the merged range
            cell.border = Border(bottom=thin_side)

        # Apply left border to the first column in the range (E4:E4)
        ws["E4"].border = Border(left=thin_side)

        # Apply right border to the last column in the range (F4:F4)
        ws["F4"].border = Border(right=thin_side)

        # Ensure the corner cells have the necessary borders
        # Top-left corner (E4) - needs both top and left borders
        ws['E4'].border = Border(left=thin_side, top=thin_side)

        # Top-right corner (F4) - needs both top and right borders
        ws['F4'].border = Border(right=thin_side, top=thin_side)

        # Since this is a single row merge, no bottom-left and bottom-right corners needed separately
        # Apply border to the outer edges of the merged range G4:I5

        # Apply top border to cells in the first row (G4:I4)
        for cell in ws["G4:I4"][0]:  # First row of the merged range
            cell.border = Border(top=thin_side)

        # Apply bottom border to cells in the last row (G5:I5)
        for cell in ws["G5:I5"][0]:  # Last row of the merged range
            cell.border = Border(bottom=thin_side)

        # Apply left border to the first column in the range (G4:G5)
        for cell in ws["G4:G5"]:  # Left column of the merged range
            cell[0].border = Border(left=thin_side)

        # Apply right border to the last column in the range (I4:I5)
        for cell in ws["I4:I5"]:  # Right column of the merged range
            cell[0].border = Border(right=thin_side)

        # Ensure the corner cells have the necessary borders
        # Top-left corner (G4) - needs both top and left borders
        ws['G4'].border = Border(left=thin_side, top=thin_side)

        # Bottom-left corner (G5) - needs both bottom and left borders
        ws['G5'].border = Border(left=thin_side, bottom=thin_side)

        # Top-right corner (I4) - needs both top and right borders
        ws['I4'].border = Border(right=thin_side, top=thin_side)

        # Bottom-right corner (I5) - needs both bottom and right borders
        ws['I5'].border = Border(right=thin_side, bottom=thin_side)

            # Define a thin border for all edges
        thin_side = Side(style='thin')
        all_borders = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

        # Apply the border to all edges of the range C5:F5
        for cell in ws["C5:F5"][0]:  # Select the first and only row in this range
            cell.border = all_borders


        # Define border styles
        medium_top_side = Side(style='medium')  # Medium top border
        thin_side = Side(style='thin')  # Thin left and right border

        # Apply medium top border to the range A40:I40
        for cell in ws["A40:I40"][0]:  # Apply to the entire row A40:I40
            # For the top row, apply only the medium top border
            cell.border = Border(top=medium_top_side)

        # Apply thin left border to A40
        ws['A40'].border = Border(left=thin_side, top=medium_top_side)

        # Apply thin right border to I40
        ws['I40'].border = Border(right=thin_side, top=medium_top_side)

    def process_excel_file(self, file_path):
        try:
            # Load data
            df = pd.read_excel(file_path, sheet_name='TIMESHEET', header=None)
            wb = load_workbook(file_path, data_only=True)

            # Create context
            context = TimesheetContext(df)

            # Extract category components
            df[['CategoryType', 'Client', 'Project']] = df.iloc[:, 0].apply(
                lambda x: pd.Series(self.extract_category_components(x))
            )

            # Drop rows where category extraction failed
            df.dropna(subset=['CategoryType', 'Client', 'Project'], inplace=True)

            # Get unique client-project pairs
            client_project_pairs = df[['Client', 'Project']].drop_duplicates().apply(tuple, axis=1).tolist()

            # Process each client-project
            for client, project in client_project_pairs:
                self.process_client_project(wb, context, df, client, project)

            # Save workbook
            wb.save(file_path)
            print(f"Processed and saved the sheets for {os.path.basename(file_path)} successfully.")

        except PermissionError:
            print(f"Error: The file {os.path.basename(file_path)} is currently open. Please close it and try again.")
        except Exception as e:
            print(f"An error occurred while processing {os.path.basename(file_path)}: {e}")

    def process_all_excel_files(self, directory):
        # List all Excel files in the directory
        excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

        # Process each Excel file in the directory
        for file_name in excel_files:
            file_path = os.path.join(directory, file_name)
            self.process_excel_file(file_path)


if __name__ == "__main__":
    # Get the current directory where the script is located
    current_directory = os.path.dirname(__file__)

    # Create an instance of TimesheetProcessor
    processor = TimesheetProcessor()

    # Process all Excel files in the current directory
    processor.process_all_excel_files(current_directory)
